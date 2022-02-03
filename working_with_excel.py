from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("apple.xlsx")
ws = wb.active

wb.create_sheet("usb")
wb.create_sheet("usbc")

ws_1 = wb["usb"]
ws_2 = wb["usbc"]

heading = ["date", "price", "adaptor", "current"]

ws_1.append(heading)
ws_2.append(heading)

for row in range(1, ws.max_row+1):
    for col in range(1, 6):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "usb":
            ws_1.append([ws["A" + str(row)].value, ws["B" + str(row)].value, ws["C" + str(row)].value,
                         ws["D" + str(row)].value])

for row in range(1, 101):
    for col in range(1, 6):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "usbc":
            ws_2.append([ws["A" + str(row)].value, ws["B" + str(row)].value, ws["C" + str(row)].value,
                         ws["D" + str(row)].value])

wb.save("apple.xlsx")
