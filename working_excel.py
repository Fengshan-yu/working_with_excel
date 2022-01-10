from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("MRI.xlsx")
ws = wb.active

wb.create_sheet("BL")
wb.create_sheet("W10")

ws_1 = wb["BL"]
ws_2 = wb["W10"]

heading = ["subject", "time point", "DeltaV", "VV"]

ws_1.append(heading)
ws_2.append(heading)

for row in range(1, 101):
    for col in range(1, 6):
        char = get_column_letter(col)
        if (ws[char + str(row)].value) == "BL":
            ws_1.append([(ws["A" + str(row)].value), (ws["B" + str(row)].value), (ws["C" + str(row)].value),
                         (ws["D" + str(row)].value)])

for row in range(1, 101):
    for col in range(1, 6):
        char = get_column_letter(col)
        if (ws[char + str(row)].value) == "W10":
            ws_2.append([(ws["A" + str(row)].value), (ws["B" + str(row)].value), (ws["C" + str(row)].value),
                         (ws["D" + str(row)].value)])

wb.save("MRI.xlsx")