from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("apple.xlsx")
ws = wb["usb"]


def get_row_data(row):
    # define a function to get data from a certain row here, so that it is easier to append data in the for loop.
    row_data = [ws["A" + str(row)].value, ws["B" + str(row)].value, ws["C" + str(row)].value, ws["D" + str(row)].value,
                ws["E" + str(row)].value, ws["F" + str(row)].value, ws["G" + str(row)].value, ws["H" + str(row)].value,
                ws["I" + str(row)].value, ws["J" + str(row)].value, ws["K" + str(row)].value, ws["L" + str(row)].value,
                ws["M" + str(row)].value, ws["N" + str(row)].value, ws["O" + str(row)].value]
    return row_data


wb.create_sheet("temp")
ws_1 = wb["temp"]

for row in range(1, ws.max_row + 1):
    for col in range(1, 16):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "white":
            ws_1.append(get_row_data(row))

wb.save("apple.xlsx")

wb = load_workbook("apple.xlsx")
ws = wb["temp"]

wb.create_sheet("temp_a")
wb.create_sheet("temp_b")
ws_2 = wb["temp_a"]
ws_3 = wb["temp_b"]

for row in range(1, ws.max_row + 1):
    for col in range(1, 16):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "12W:
            ws_2.append(get_row_data(row))

for row in range(1, ws.max_row + 1):
    for col in range(1, 16):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "20W":
            ws_3.append(get_row_data(row))

del wb["temp"]
wb.save("apple.xlsx")

wb = load_workbook("apple.xlsx")
ws = wb["temp_a"]

wb.create_sheet("sorted_apple")
ws_4 = wb["sorted_apple"]
for row in range(1, ws.max_row + 1):
    for col in range(1, 16):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "white":
            ws_4.append(get_row_data(row))

for row in range(1, ws.max_row + 1):
    for col in range(1, 16):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "black":
            ws_4.append(get_row_data(row))

del wb["temp_a"]
wb.save("apple.xlsx")

wb = load_workbook("apple.xlsx")
ws = wb["temp_b"]
ws_4 = wb["sorted_apple"]

for row in range(1, ws.max_row + 1):
    for col in range(1, 16):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "white":
            ws_4.append(get_row_data(row))

for row in range(1, ws.max_row + 1):
    for col in range(1, 16):
        char = get_column_letter(col)
        if ws[char + str(row)].value == "black":
            ws_4.append(get_row_data(row))

del wb["temp_b"]
wb.save("apple.xlsx")
